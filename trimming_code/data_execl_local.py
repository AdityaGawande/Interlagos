from openpyxl import Workbook, load_workbook
import os

def write_to_excel_cell(file_path, sheet_name, cell_address, value):
    # Load workbook if it exists; otherwise, create a new one
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    # Use existing sheet or create it
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Write the value to the specified cell
    ws[cell_address] = value

    # Save the workbook
    wb.save(file_path)
    print(f"Written '{value}' to {sheet_name}:{cell_address} in '{file_path}'")
